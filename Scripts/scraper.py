#This program opens, edits, and saves an excel file used to track a database of information based on Discharge Review Board data

#imports for excel file manipulation with openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

#Search text files and retreive information stored as tuples/matrices
import drbsearch


#Write information from tuplles/matrices to excel file
dest_filename = 'excel.xlsx'
wb = load_workbook(dest_filename)
#This program opens, edits, and saves an excel file used to track a database of information based on Discharge Review Board data

#imports for excel file manipulation with openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

#Search text files and retreive information stored as tuples/matrices
import drbsearch


#Write information from lists to excel file
dest_filename = 'excel.xlsx'
wb = load_workbook(dest_filename)

ws1 = wb.active

for row in range(2,len(drbsearch.onlyfiles)):
    ws1.cell(column=1, row=row, value=drbsearch.onlyfiles[row-2])
for row in range(2,len(drbsearch.onlygrants)):
    ws1.cell(column=2, row=row, value=drbsearch.onlygrants[row-2])
for row in range(2,len(drbsearch.onlydenies)):
    ws1.cell(column=3, row=row, value=drbsearch.onlydenies[row-2])
for row in range(2,len(drbsearch.onlymentalhealth)):
    ws1.cell(column=4, row=row, value=drbsearch.onlymentalhealth[row-2])

wb.save(dest_filename)

ws1 = wb.active

for row in range(2,len(drbsearch.onlyfiles)):
    ws1.cell(column=1, row=row, value=drbsearch.onlyfiles[row-2])
for row in range(2,len(drbsearch.onlygrants)):
    ws1.cell(column=2, row=row, value=drbsearch.onlygrants[row-2])
for row in range(2,len(drbsearch.onlydenies)):
    ws1.cell(column=3, row=row, value=drbsearch.onlydenies[row-2])
for row in range(2,len(drbsearch.onlymentalhealth)):
    ws1.cell(column=4, row=row, value=drbsearch.onlymentalhealth[row-2])

wb.save(dest_filename)
